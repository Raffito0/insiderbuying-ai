"""
Generate 3 professional portfolio Excel samples for a data specialist freelancer.
Each file has realistic fake data, professional formatting, and summary sheets.
"""

import random
import string
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Shared formatting helpers
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

LIGHT_GRAY_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")


def style_header(ws, header_color, num_cols):
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, num_rows, num_cols, start_row=2):
    for row in range(start_row, start_row + num_rows):
        fill = LIGHT_GRAY_FILL if row % 2 == 0 else WHITE_FILL
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def auto_fit_columns(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def random_date(start, end):
    delta = end - start
    rand_days = random.randint(0, delta.days)
    rand_secs = random.randint(0, 86399)
    return start + timedelta(days=rand_days, seconds=rand_secs)


def style_summary_sheet(ws, header_color):
    """Apply professional formatting to a summary/dashboard sheet."""
    title_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=13)
    label_font = Font(bold=True, size=11, color="333333")
    value_font = Font(size=11, color="111111")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Title row (row 1)
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = title_fill
        c.font = title_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Label / value styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for i, cell in enumerate(row):
            if i == 0:
                cell.font = label_font
            else:
                cell.font = value_font

    auto_fit_columns(ws, min_width=18, max_width=45)


# ===========================================================================
# SAMPLE 1: E-commerce Product Scraping
# ===========================================================================

def generate_ecommerce():
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Data"

    # --- Realistic product catalog ---
    products_pool = [
        ("Samsung Galaxy S24 Ultra 256GB", "Samsung", "Smartphones", 1199.99),
        ("Samsung Galaxy S24+ 128GB", "Samsung", "Smartphones", 999.99),
        ("Samsung Galaxy A55 5G", "Samsung", "Smartphones", 429.99),
        ("Apple iPhone 15 Pro Max 256GB", "Apple", "Smartphones", 1199.00),
        ("Apple iPhone 15 128GB", "Apple", "Smartphones", 799.00),
        ("Apple iPhone SE (2022) 64GB", "Apple", "Smartphones", 429.00),
        ("Google Pixel 8 Pro 128GB", "Google", "Smartphones", 999.00),
        ("Google Pixel 8a 128GB", "Google", "Smartphones", 499.00),
        ("OnePlus 12 256GB", "OnePlus", "Smartphones", 799.99),
        ("Sony WH-1000XM5 Wireless Headphones", "Sony", "Audio", 349.99),
        ("Sony WH-1000XM4 Wireless Headphones", "Sony", "Audio", 248.00),
        ("Apple AirPods Pro 2nd Gen USB-C", "Apple", "Audio", 249.00),
        ("Apple AirPods Max Silver", "Apple", "Audio", 549.00),
        ("Bose QuietComfort Ultra Headphones", "Bose", "Audio", 429.00),
        ("Bose QuietComfort 45 Headphones", "Bose", "Audio", 279.00),
        ("JBL Charge 5 Portable Speaker", "JBL", "Audio", 179.95),
        ("JBL Flip 6 Bluetooth Speaker", "JBL", "Audio", 129.95),
        ("Sennheiser Momentum 4 Wireless", "Sennheiser", "Audio", 349.95),
        ("Sony Alpha a7 IV Mirrorless Camera Body", "Sony", "Cameras", 2498.00),
        ("Sony Alpha a6700 Mirrorless Camera", "Sony", "Cameras", 1398.00),
        ("Canon EOS R6 Mark II Body", "Canon", "Cameras", 2499.00),
        ("Canon EOS R50 Mirrorless Camera", "Canon", "Cameras", 679.00),
        ("Nikon Z8 Mirrorless Camera Body", "Nikon", "Cameras", 3996.95),
        ("Nikon Z fc Body DX-Format", "Nikon", "Cameras", 856.95),
        ("Fujifilm X-T5 Mirrorless Camera Body", "Fujifilm", "Cameras", 1699.00),
        ("GoPro HERO12 Black", "GoPro", "Cameras", 349.99),
        ("DJI Mini 4 Pro Drone", "DJI", "Cameras", 759.00),
        ("DJI Air 3 Fly More Combo", "DJI", "Cameras", 1349.00),
        ("Apple MacBook Pro 14\" M3 Pro 18GB", "Apple", "Computing", 1999.00),
        ("Apple MacBook Air 15\" M3 16GB", "Apple", "Computing", 1299.00),
        ("Apple MacBook Air 13\" M3 8GB", "Apple", "Computing", 1099.00),
        ("Dell XPS 15 9530 i7 16GB", "Dell", "Computing", 1499.99),
        ("Dell XPS 13 Plus i7 16GB", "Dell", "Computing", 1299.99),
        ("Lenovo ThinkPad X1 Carbon Gen 11", "Lenovo", "Computing", 1429.00),
        ("Lenovo IdeaPad 5 15\" Ryzen 7", "Lenovo", "Computing", 649.99),
        ("HP Spectre x360 14 i7 16GB", "HP", "Computing", 1449.99),
        ("HP Pavilion 15 Ryzen 5 8GB", "HP", "Computing", 579.99),
        ("ASUS ROG Zephyrus G14 RTX 4060", "ASUS", "Computing", 1599.99),
        ("ASUS Zenbook 14 OLED UX3405", "ASUS", "Computing", 1099.99),
        ("Apple iPad Pro 12.9\" M2 256GB", "Apple", "Tablets", 1099.00),
        ("Apple iPad Air 11\" M2 128GB", "Apple", "Tablets", 599.00),
        ("Apple iPad 10th Gen 64GB", "Apple", "Tablets", 349.00),
        ("Samsung Galaxy Tab S9 Ultra 256GB", "Samsung", "Tablets", 1199.99),
        ("Samsung Galaxy Tab S9 FE 128GB", "Samsung", "Tablets", 449.99),
        ("Amazon Fire HD 10 (2023) 32GB", "Amazon", "Tablets", 139.99),
        ("Apple Watch Ultra 2", "Apple", "Wearables", 799.00),
        ("Apple Watch Series 9 45mm", "Apple", "Wearables", 429.00),
        ("Samsung Galaxy Watch 6 Classic 47mm", "Samsung", "Wearables", 399.99),
        ("Garmin Fenix 7X Solar", "Garmin", "Wearables", 799.99),
        ("Fitbit Charge 6", "Fitbit", "Wearables", 159.95),
        ("Sony PlayStation 5 Slim Digital", "Sony", "Gaming", 449.99),
        ("Sony PlayStation 5 Slim Disc", "Sony", "Gaming", 499.99),
        ("Microsoft Xbox Series X 1TB", "Microsoft", "Gaming", 499.99),
        ("Microsoft Xbox Series S 512GB", "Microsoft", "Gaming", 299.99),
        ("Nintendo Switch OLED Model", "Nintendo", "Gaming", 349.99),
        ("Meta Quest 3 128GB", "Meta", "Gaming", 499.99),
        ("Logitech MX Master 3S Wireless Mouse", "Logitech", "Accessories", 99.99),
        ("Logitech MX Keys S Keyboard", "Logitech", "Accessories", 109.99),
        ("Apple Magic Keyboard with Touch ID", "Apple", "Accessories", 199.00),
        ("Anker 737 Power Bank 24000mAh", "Anker", "Accessories", 109.99),
        ("Anker Nano II 65W GaN Charger", "Anker", "Accessories", 35.99),
        ("Samsung T7 Shield 2TB Portable SSD", "Samsung", "Storage", 159.99),
        ("Samsung 990 Pro 2TB NVMe SSD", "Samsung", "Storage", 179.99),
        ("WD Black SN850X 2TB NVMe SSD", "Western Digital", "Storage", 149.99),
        ("Seagate Expansion 5TB External HDD", "Seagate", "Storage", 119.99),
        ("SanDisk Extreme Pro 1TB microSD", "SanDisk", "Storage", 109.99),
        ("LG C3 65\" OLED 4K Smart TV", "LG", "TVs & Displays", 1496.99),
        ("LG C3 55\" OLED 4K Smart TV", "LG", "TVs & Displays", 1096.99),
        ("Samsung S90C 65\" OLED 4K Smart TV", "Samsung", "TVs & Displays", 1597.99),
        ("Sony A80L 65\" OLED 4K Smart TV", "Sony", "TVs & Displays", 1698.00),
        ("TCL 65\" Class S4 4K Smart TV", "TCL", "TVs & Displays", 379.99),
        ("Hisense 55\" U8K Mini-LED 4K TV", "Hisense", "TVs & Displays", 749.99),
        ("Dell UltraSharp U2723QE 27\" 4K Monitor", "Dell", "TVs & Displays", 519.99),
        ("ASUS ProArt PA278QV 27\" Monitor", "ASUS", "TVs & Displays", 309.00),
        ("LG 27GP850-B 27\" Gaming Monitor", "LG", "TVs & Displays", 349.99),
        ("Dyson V15 Detect Absolute Vacuum", "Dyson", "Home", 749.99),
        ("Dyson Purifier Cool TP07", "Dyson", "Home", 569.99),
        ("iRobot Roomba j7+ Self-Emptying Robot", "iRobot", "Home", 599.99),
        ("Sonos Era 300 Speaker", "Sonos", "Audio", 449.00),
        ("Sonos Beam Gen 2 Soundbar", "Sonos", "Audio", 449.00),
    ]

    headers = [
        "Product Name", "Brand", "Price (USD)", "Original Price (USD)",
        "Discount %", "Rating", "Reviews", "Availability", "Category",
        "Product URL", "SKU", "Last Scraped"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    style_header(ws, "1A1A2E", len(headers))

    scrape_start = datetime(2026, 3, 1)
    scrape_end = datetime(2026, 3, 23)
    availability_weights = ["In Stock"] * 78 + ["Out of Stock"] * 12 + ["Limited Stock"] * 10

    rows_data = []
    # Use each product at least twice, fill to 200
    expanded = list(products_pool) * 3
    random.shuffle(expanded)
    selected = expanded[:200]

    for i, (name, brand, category, base_price) in enumerate(selected, start=1):
        # Price variation
        price = round(base_price * random.uniform(0.85, 1.05), 2)
        is_sale = random.random() < 0.40
        if is_sale:
            discount_pct = random.choice([5, 8, 10, 12, 15, 18, 20, 22, 25, 30])
            orig_price = round(price / (1 - discount_pct / 100), 2)
        else:
            orig_price = price
            discount_pct = 0

        rating = round(random.triangular(2.5, 5.0, 4.3), 1)
        rating = min(5.0, max(1.0, rating))
        reviews = int(random.triangular(10, 15000, 450))
        avail = random.choice(availability_weights)
        slug = name.lower().replace(" ", "-").replace('"', "").replace("'", "")
        slug = slug.replace("/", "-").replace("(", "").replace(")", "")
        url = f"/products/{slug}"
        prefix = "".join(random.choices(string.ascii_uppercase, k=3))
        sku = f"{prefix}-{random.randint(10000, 99999)}"
        scraped = random_date(scrape_start, scrape_end)

        row = [
            name, brand, price, orig_price, discount_pct,
            rating, reviews, avail, category, url, sku,
            scraped.strftime("%Y-%m-%d %H:%M")
        ]
        rows_data.append(row)

    for r_idx, row in enumerate(rows_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    style_data_rows(ws, len(rows_data), len(headers))

    # Currency formatting
    for row in range(2, len(rows_data) + 2):
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).number_format = '0"%"'
        # Rating conditional coloring
        rating_val = ws.cell(row=row, column=6).value
        if rating_val and rating_val >= 4.0:
            ws.cell(row=row, column=6).fill = GREEN_FILL
            ws.cell(row=row, column=6).font = GREEN_FONT
        elif rating_val and rating_val >= 3.0:
            ws.cell(row=row, column=6).fill = YELLOW_FILL
            ws.cell(row=row, column=6).font = YELLOW_FONT
        elif rating_val:
            ws.cell(row=row, column=6).fill = RED_FILL
            ws.cell(row=row, column=6).font = RED_FONT

        ws.cell(row=row, column=7).number_format = '#,##0'

    auto_fit_columns(ws)
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.append(["E-Commerce Scraping Summary", "", ""])
    ws2.merge_cells("A1:C1")

    prices = [r[2] for r in rows_data]
    ratings = [r[5] for r in rows_data]
    categories = {}
    for r in rows_data:
        categories[r[8]] = categories.get(r[8], 0) + 1

    summary_rows = [
        ["Total Products Scraped", len(rows_data), ""],
        ["Average Price", f"${sum(prices)/len(prices):.2f}", ""],
        ["Price Range", f"${min(prices):.2f} - ${max(prices):.2f}", ""],
        ["Average Rating", f"{sum(ratings)/len(ratings):.1f}", ""],
        ["Products on Sale", sum(1 for r in rows_data if r[4] > 0), ""],
        ["In Stock", sum(1 for r in rows_data if r[7] == "In Stock"), ""],
        ["Out of Stock", sum(1 for r in rows_data if r[7] == "Out of Stock"), ""],
        ["Limited Stock", sum(1 for r in rows_data if r[7] == "Limited Stock"), ""],
        ["", "", ""],
        ["Category", "Count", "Avg Price"],
    ]
    for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
        cat_prices = [r[2] for r in rows_data if r[8] == cat]
        summary_rows.append([cat, cnt, f"${sum(cat_prices)/len(cat_prices):.2f}"])

    for row in summary_rows:
        ws2.append(row)

    style_summary_sheet(ws2, "1A1A2E")

    return wb


# ===========================================================================
# SAMPLE 2: B2B Dental Leads - Texas
# ===========================================================================

def generate_dental_leads():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dental Leads TX"

    # --- Realistic name/clinic pools ---
    first_names_m = [
        "James", "Robert", "Michael", "David", "William", "Richard", "Joseph",
        "Thomas", "Christopher", "Daniel", "Matthew", "Anthony", "Mark", "Steven",
        "Andrew", "Paul", "Joshua", "Kenneth", "Kevin", "Brian", "Jason", "Ryan",
        "Eric", "Stephen", "Timothy", "Jeffrey", "Nathan", "Scott", "Benjamin",
        "Patrick", "Carlos", "Miguel", "Alejandro", "Luis", "Rafael",
    ]
    first_names_f = [
        "Mary", "Patricia", "Jennifer", "Linda", "Barbara", "Elizabeth", "Susan",
        "Jessica", "Sarah", "Karen", "Lisa", "Nancy", "Margaret", "Sandra",
        "Ashley", "Dorothy", "Kimberly", "Emily", "Donna", "Michelle", "Carol",
        "Amanda", "Melissa", "Deborah", "Stephanie", "Rebecca", "Sharon",
        "Laura", "Christina", "Maria", "Ana", "Gabriela", "Sofia",
    ]
    last_names = [
        "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
        "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
        "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
        "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark",
        "Ramirez", "Lewis", "Robinson", "Walker", "Young", "Allen", "King",
        "Wright", "Scott", "Torres", "Nguyen", "Hill", "Flores", "Green",
        "Adams", "Nelson", "Baker", "Hall", "Rivera", "Campbell", "Mitchell",
        "Carter", "Roberts", "Patel", "Shah", "Gupta", "Chen", "Park",
    ]

    descriptors = [
        "Bright Smile", "Premier", "Advanced", "Family", "Gentle Care",
        "Elite", "Comfort", "Modern", "Professional", "Precision",
        "Total Care", "Complete", "Affordable", "Quality", "Friendly",
        "Caring", "Creative", "Classic", "Heritage", "Signature",
        "Radiant", "Legacy", "Harmony", "Excellence", "Peak",
        "Crossroads", "Lakeside", "Meadow", "Riverside", "Valley",
        "Sunrise", "Sunset", "Hilltop", "Garden", "Cedar Park",
        "Live Oak", "Pecan Grove", "Bluebonnet", "Magnolia", "Lone Star",
        "Southern", "Texan", "Republic", "Frontier", "Ranch Road",
    ]
    suffixes = [
        "Dental", "Dental Care", "Dentistry", "Dental Group",
        "Dental Arts", "Dental Associates", "Dental Studio",
        "Family Dentistry", "Dental Clinic", "Dental Center",
        "Dental Practice", "Dental Wellness", "Oral Health",
    ]

    cities_zips = [
        ("Austin", ["73301", "78701", "78702", "78704", "78731", "78741", "78745", "78748", "78749", "78750", "78757", "78759"]),
        ("Houston", ["77001", "77002", "77003", "77004", "77005", "77006", "77007", "77008", "77019", "77024", "77025", "77027", "77030", "77056", "77057", "77063", "77077", "77079", "77082", "77084"]),
        ("Dallas", ["75201", "75202", "75204", "75205", "75206", "75207", "75208", "75209", "75210", "75214", "75219", "75225", "75226", "75228", "75230", "75231", "75235", "75240", "75243", "75248"]),
        ("San Antonio", ["78201", "78202", "78204", "78205", "78207", "78209", "78210", "78212", "78213", "78215", "78216", "78217", "78218", "78220", "78223", "78224", "78227", "78229", "78230", "78232"]),
        ("Fort Worth", ["76102", "76103", "76104", "76105", "76106", "76107", "76109", "76110", "76112", "76116", "76120", "76123", "76126", "76129", "76132", "76133", "76134", "76137", "76140", "76148"]),
        ("Plano", ["75023", "75024", "75025", "75074", "75075", "75093"]),
        ("El Paso", ["79901", "79902", "79903", "79904", "79905", "79906", "79907", "79912", "79915", "79922", "79924", "79925", "79927", "79930", "79932", "79934", "79935", "79936"]),
        ("Arlington", ["76001", "76002", "76006", "76010", "76011", "76012", "76013", "76014", "76015", "76016", "76017", "76018"]),
        ("Frisco", ["75033", "75034", "75035"]),
        ("McKinney", ["75069", "75070", "75071"]),
        ("Round Rock", ["78664", "78665", "78681"]),
        ("Irving", ["75038", "75039", "75060", "75061", "75062", "75063"]),
        ("Laredo", ["78040", "78041", "78043", "78045", "78046"]),
        ("Lubbock", ["79401", "79403", "79404", "79407", "79410", "79411", "79412", "79413", "79414", "79415", "79416"]),
        ("Corpus Christi", ["78401", "78404", "78405", "78408", "78410", "78411", "78412", "78413", "78414", "78415", "78416", "78418"]),
        ("Sugar Land", ["77478", "77479", "77498"]),
        ("The Woodlands", ["77380", "77381", "77382", "77384", "77385", "77386", "77389"]),
        ("Katy", ["77449", "77450", "77493", "77494"]),
        ("Cedar Park", ["78613"]),
        ("Flower Mound", ["75022", "75028"]),
        ("Pearland", ["77581", "77584", "77588"]),
        ("Midland", ["79701", "79703", "79705", "79706", "79707"]),
        ("Odessa", ["79761", "79762", "79763", "79764", "79765"]),
        ("Denton", ["76201", "76205", "76207", "76208", "76209", "76210"]),
        ("Waco", ["76701", "76704", "76706", "76707", "76708", "76710", "76711", "76712"]),
        ("Killeen", ["76541", "76542", "76543", "76544", "76549"]),
        ("Tyler", ["75701", "75702", "75703", "75707", "75708", "75709"]),
        ("Beaumont", ["77701", "77702", "77703", "77706", "77707", "77708"]),
        ("College Station", ["77840", "77845"]),
        ("San Marcos", ["78666"]),
        ("Georgetown", ["78626", "78628", "78633"]),
    ]

    city_area_codes = {
        "Austin": ["512"], "Round Rock": ["512"], "Cedar Park": ["512"],
        "San Marcos": ["512"], "Georgetown": ["512"],
        "Houston": ["713", "281", "832"], "Sugar Land": ["281", "832"],
        "The Woodlands": ["281", "832"], "Katy": ["281", "832"],
        "Pearland": ["281", "832"],
        "Dallas": ["214", "469", "972"], "Plano": ["214", "469", "972"],
        "Frisco": ["214", "469", "972"], "McKinney": ["214", "469", "972"],
        "Irving": ["214", "469", "972"], "Flower Mound": ["214", "469", "972"],
        "Denton": ["940"],
        "San Antonio": ["210"],
        "Fort Worth": ["817"], "Arlington": ["817"],
        "El Paso": ["915"],
        "Laredo": ["956"],
        "Lubbock": ["806"],
        "Corpus Christi": ["361"],
        "Midland": ["432"], "Odessa": ["432"],
        "Waco": ["254"], "Killeen": ["254"],
        "Tyler": ["903"],
        "Beaumont": ["409"],
        "College Station": ["979"],
    }

    street_names = [
        "Main St", "Commerce St", "Elm St", "Oak Dr", "Maple Ave",
        "Congress Ave", "Guadalupe St", "Westheimer Rd", "Memorial Dr",
        "Preston Rd", "Greenville Ave", "McKinney Ave", "Travis St",
        "Lamar Blvd", "Burnet Rd", "Anderson Ln", "Parmer Ln",
        "FM 1960", "Highway 290", "Loop 360", "Research Blvd",
        "Bee Cave Rd", "Slaughter Ln", "Brodie Ln", "Manchaca Rd",
        "Huebner Rd", "Bandera Rd", "Fredericksburg Rd", "Broadway St",
        "Nacogdoches Rd", "Wurzbach Rd", "Medical Dr", "Voss Rd",
        "Kirby Dr", "Montrose Blvd", "Richmond Ave", "Bissonnet St",
        "Bellaire Blvd", "Gessner Rd", "Fondren Rd", "Hillcroft Ave",
        "Harry Hines Blvd", "Northwest Hwy", "Mockingbird Ln",
        "Lovers Ln", "Inwood Rd", "Lemmon Ave", "Cedar Springs Rd",
        "Camp Bowie Blvd", "University Dr", "Hulen St", "Bryant Irvin Rd",
    ]

    titles = (
        ["Owner / DDS"] * 25 + ["DDS"] * 20 + ["DMD"] * 15 +
        ["Office Manager"] * 20 + ["Marketing Director"] * 8 +
        ["Practice Administrator"] * 7 + ["Associate Dentist"] * 5
    )
    sources = (
        ["Google Maps"] * 35 + ["Yelp"] * 20 + ["ADA Directory"] * 15 +
        ["ZoomInfo"] * 10 + ["LinkedIn"] * 10 + ["Healthgrades"] * 5 +
        ["Yellow Pages"] * 5
    )

    headers = [
        "Company Name", "Contact Name", "Title", "Email", "Phone",
        "Website", "Address", "City", "State", "Zip Code",
        "Employee Count", "Estimated Revenue", "Email Verified",
        "LinkedIn URL", "Source", "Date Collected"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, "0D5016", len(headers))

    used_clinic_names = set()
    rows_data = []
    collect_start = datetime(2026, 3, 1)
    collect_end = datetime(2026, 3, 23)

    for i in range(150):
        # Generate unique clinic name
        while True:
            desc = random.choice(descriptors)
            suf = random.choice(suffixes)
            clinic_name = f"{desc} {suf}"
            if clinic_name not in used_clinic_names:
                used_clinic_names.add(clinic_name)
                break

        # Sometimes use "Dr. LastName Dentistry" format
        if random.random() < 0.20:
            ln = random.choice(last_names)
            suf2 = random.choice(["Dentistry", "Dental", "Dental Care", "DDS"])
            clinic_name = f"Dr. {ln} {suf2}"
            if clinic_name in used_clinic_names:
                clinic_name = f"{desc} {suf}"
            used_clinic_names.add(clinic_name)

        is_male = random.random() < 0.55
        first = random.choice(first_names_m if is_male else first_names_f)
        last = random.choice(last_names)
        contact = f"{first} {last}"
        title = random.choice(titles)

        city, zips = random.choice(cities_zips)
        zipcode = random.choice(zips)
        area_codes = city_area_codes.get(city, ["512"])
        ac = random.choice(area_codes)
        phone = f"({ac}) {random.randint(200,999)}-{random.randint(1000,9999)}"

        domain_base = clinic_name.lower().replace("dr. ", "").replace(" ", "").replace("'", "").replace("&", "and")
        domain_base = domain_base[:20]
        domain = f"www.{domain_base}.com"

        if random.random() < 0.6:
            email = f"{first.lower()}.{last.lower()}@{domain_base}.com"
        elif random.random() < 0.5:
            email = f"{first.lower()[0]}{last.lower()}@{domain_base}.com"
        else:
            email = f"info@{domain_base}.com"

        street_num = random.randint(100, 19999)
        street = random.choice(street_names)
        suite = f", Suite {random.randint(100, 999)}" if random.random() < 0.45 else ""
        address = f"{street_num} {street}{suite}"

        emp_count = random.choice(
            list(range(2, 8)) * 4 + list(range(8, 20)) * 3 + list(range(20, 51))
        )
        base_rev = emp_count * random.randint(40000, 120000)
        revenue = max(200000, min(5000000, base_rev))

        verified = "Yes" if random.random() < 0.95 else "No"

        li_slug = f"{first.lower()}-{last.lower()}-{random.randint(100, 999)}{''.join(random.choices('abcdef', k=3))}"
        linkedin = f"https://linkedin.com/in/{li_slug}"

        source = random.choice(sources)
        collected = random_date(collect_start, collect_end).strftime("%Y-%m-%d")

        rows_data.append([
            clinic_name, contact, title, email, phone,
            domain, address, city, "TX", zipcode,
            emp_count, revenue, verified, linkedin, source, collected
        ])

    for r_idx, row in enumerate(rows_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    style_data_rows(ws, len(rows_data), len(headers))

    # Formatting
    for row in range(2, len(rows_data) + 2):
        ws.cell(row=row, column=12).number_format = '$#,##0'
        # Email Verified coloring
        v = ws.cell(row=row, column=13).value
        if v == "Yes":
            ws.cell(row=row, column=13).fill = GREEN_FILL
            ws.cell(row=row, column=13).font = GREEN_FONT
        else:
            ws.cell(row=row, column=13).fill = RED_FILL
            ws.cell(row=row, column=13).font = RED_FONT

    auto_fit_columns(ws)
    ws.freeze_panes = "A2"

    # --- Verification Summary sheet ---
    ws2 = wb.create_sheet("Verification Summary")
    ws2.append(["B2B Lead Verification Summary", "", ""])
    ws2.merge_cells("A1:C1")

    verified_count = sum(1 for r in rows_data if r[12] == "Yes")
    source_dist = {}
    city_dist = {}
    for r in rows_data:
        source_dist[r[14]] = source_dist.get(r[14], 0) + 1
        city_dist[r[7]] = city_dist.get(r[7], 0) + 1

    ws2.append(["Total Leads", len(rows_data), ""])
    ws2.append(["Email Verified", f"{verified_count} ({verified_count/len(rows_data)*100:.1f}%)", ""])
    ws2.append(["Not Verified", f"{len(rows_data)-verified_count} ({(len(rows_data)-verified_count)/len(rows_data)*100:.1f}%)", ""])
    ws2.append(["", "", ""])
    ws2.append(["Source", "Count", "% of Total"])
    for s, c in sorted(source_dist.items(), key=lambda x: -x[1]):
        ws2.append([s, c, f"{c/len(rows_data)*100:.1f}%"])
    ws2.append(["", "", ""])
    ws2.append(["City", "Count", "% of Total"])
    for city, c in sorted(city_dist.items(), key=lambda x: -x[1])[:15]:
        ws2.append([city, c, f"{c/len(rows_data)*100:.1f}%"])
    if len(city_dist) > 15:
        rest = sum(c for city, c in sorted(city_dist.items(), key=lambda x: -x[1])[15:])
        ws2.append(["Other cities", rest, f"{rest/len(rows_data)*100:.1f}%"])

    style_summary_sheet(ws2, "0D5016")

    return wb


# ===========================================================================
# SAMPLE 3: Invoice Processing (PDF to Excel)
# ===========================================================================

def generate_invoice_processing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Invoices"

    vendor_pool = [
        "Apex Supply Co.", "MetroTech Solutions", "Pacific Coast Logistics",
        "Clearwater Systems Inc.", "Summit Digital Group", "Horizon Cloud Services",
        "Brightline Marketing", "Cascade Manufacturing", "Atlas Engineering Corp.",
        "NovaStar Technologies", "Pinnacle Staffing Solutions", "Redwood Analytics",
        "Sterling Office Products", "Vanguard IT Consulting", "Ironclad Security Systems",
        "BlueShore Cleaning Services", "Keystone Printing & Design", "Northwind Supplies",
        "Quantum Data Corp.", "Trident Logistics LLC", "Golden Gate Catering",
        "Eagle Eye Surveillance", "OmniTech Solutions", "Westfield Legal Services",
        "Cornerstone Insurance Group", "Granite Construction Co.", "Silverline Communications",
        "Harbor Point Financial", "Crestview Property Management", "Evergreen Landscaping",
        "Velocity Courier Services", "Spark Creative Agency", "Bridgeport Engineering",
        "Lakeview Medical Supplies", "Phoenix Energy Solutions", "Diamond Data Recovery",
        "Oakwood Furniture Co.", "Prism Software Inc.", "FrostByte IT Services",
        "Crimson Marketing Group",
    ]

    line_item_pool = [
        "Web hosting - Annual", "Office supplies Q1 2026", "Consulting services - March",
        "Software license renewal (12 months)", "Cloud storage - 5TB plan",
        "Monthly maintenance contract", "Network equipment - Cisco switches",
        "Professional development training", "Print services - March brochures",
        "IT support retainer - Q1", "Janitorial services - March",
        "Marketing campaign management", "Legal consultation (10 hrs)",
        "Graphic design services - Logo redesign", "Security system monitoring - Q1",
        "Catering - Q1 executive meetings", "Courier services - March",
        "Accounting services - Monthly", "Insurance premium - Q1",
        "Copier lease - Monthly", "Phone system maintenance",
        "SEO optimization package", "Data backup service - Monthly",
        "Employee background checks (batch)", "HVAC maintenance - Quarterly",
        "Landscaping services - March", "Window cleaning - Quarterly",
        "Fire safety inspection", "Elevator maintenance - Monthly",
        "Pest control - Quarterly", "Water cooler rental - Monthly",
        "Shredding services - Monthly", "Parking lot maintenance",
        "Building security patrol - Monthly", "Waste management - Monthly",
        "Electrical work - Office renovation Phase 2", "Plumbing repair - 3rd floor",
        "Carpet cleaning - Main office", "Furniture delivery and assembly",
        "Server rack installation", "Fiber optic cable installation",
        "UPS battery replacement (12 units)", "Conference room AV setup",
        "Desk chairs (24 units)", "Standing desks (8 units)",
        "Breakroom appliances", "First aid supplies replenishment",
        "Safety equipment - Annual", "Forklift certification training",
        "Commercial vehicle maintenance", "Fuel cards - March fleet",
    ]

    notes_pool = [
        "", "", "", "", "", "", "",  # most blank
        "Early payment discount applied (2%)",
        "Recurring monthly charge",
        "Approved by Finance - J. Williams",
        "Net 60 terms per agreement",
        "Final payment of 3-part installment",
        "Rush delivery surcharge included",
        "Contract renewal - Year 3 of 5",
        "Adjusted per change order #14",
        "Tax exempt - resale certificate on file",
        "Split billing: 60% Dept A / 40% Dept B",
        "Includes after-hours labor premium",
        "Quarterly review scheduled for April",
    ]

    headers = [
        "Invoice Number", "Vendor Name", "Invoice Date", "Due Date",
        "Status", "Line Items Description", "Quantity", "Unit Price",
        "Total Amount", "Tax Amount (8.25%)", "Grand Total",
        "Payment Method", "PO Number", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, "C44900", len(headers))

    status_weights = ["Paid"] * 55 + ["Pending"] * 30 + ["Overdue"] * 15
    payment_methods = (
        ["Wire Transfer"] * 25 + ["ACH"] * 30 + ["Check"] * 20 +
        ["Credit Card"] * 15 + ["Corporate Card"] * 10
    )

    rows_data = []
    inv_start = datetime(2026, 1, 5)
    inv_end = datetime(2026, 3, 22)

    for i in range(120):
        inv_num = f"INV-2026-{i+1:04d}"
        vendor = random.choice(vendor_pool)
        inv_date = random_date(inv_start, inv_end)
        net_days = random.choice([30, 30, 30, 45, 45, 60])
        due_date = inv_date + timedelta(days=net_days)

        status = random.choice(status_weights)
        # Make status realistic: overdue only if due_date < today
        today = datetime(2026, 3, 23)
        if status == "Overdue" and due_date >= today:
            status = "Pending"
        if status == "Paid" and due_date > today and random.random() < 0.3:
            status = "Pending"  # some future invoices still pending

        desc = random.choice(line_item_pool)
        quantity = random.choice(
            [1] * 40 + [2] * 15 + [3] * 10 + [5] * 8 +
            [10] * 7 + [12] * 5 + [24] * 5 + [50] * 3 +
            list(range(6, 100))
        )
        # Unit price based on description type
        if "Monthly" in desc or "monthly" in desc:
            unit_price = round(random.uniform(75, 850), 2)
        elif "Annual" in desc or "12 months" in desc:
            unit_price = round(random.uniform(200, 5000), 2)
        elif "units" in desc:
            unit_price = round(random.uniform(45, 800), 2)
        elif "Q1" in desc or "Quarterly" in desc:
            unit_price = round(random.uniform(150, 2500), 2)
        else:
            unit_price = round(random.uniform(25, 3500), 2)

        total = round(quantity * unit_price, 2)
        tax = round(total * 0.0825, 2)
        grand_total = round(total + tax, 2)

        payment = random.choice(payment_methods)
        has_po = random.random() < 0.70
        po_num = f"PO-{random.randint(10000, 99999)}" if has_po else ""
        note = random.choice(notes_pool)

        rows_data.append([
            inv_num, vendor, inv_date.strftime("%Y-%m-%d"),
            due_date.strftime("%Y-%m-%d"), status, desc,
            quantity, unit_price, total, tax, grand_total,
            payment, po_num, note
        ])

    for r_idx, row in enumerate(rows_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    style_data_rows(ws, len(rows_data), len(headers))

    # Formatting
    for row in range(2, len(rows_data) + 2):
        for money_col in [8, 9, 10, 11]:
            ws.cell(row=row, column=money_col).number_format = '$#,##0.00'
        # Status coloring
        st = ws.cell(row=row, column=5).value
        if st == "Paid":
            ws.cell(row=row, column=5).fill = GREEN_FILL
            ws.cell(row=row, column=5).font = GREEN_FONT
        elif st == "Pending":
            ws.cell(row=row, column=5).fill = YELLOW_FILL
            ws.cell(row=row, column=5).font = YELLOW_FONT
        elif st == "Overdue":
            ws.cell(row=row, column=5).fill = RED_FILL
            ws.cell(row=row, column=5).font = RED_FONT

    # Grand Total SUM row
    sum_row = len(rows_data) + 2
    sum_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    sum_font = Font(bold=True, size=11, color="C44900")
    ws.cell(row=sum_row, column=1, value="TOTALS").font = sum_font
    ws.cell(row=sum_row, column=1).fill = sum_fill
    for col in range(2, len(headers) + 1):
        ws.cell(row=sum_row, column=col).fill = sum_fill
        ws.cell(row=sum_row, column=col).border = THIN_BORDER

    for col_idx, col_letter in [(9, "I"), (10, "J"), (11, "K")]:
        ws.cell(row=sum_row, column=col_idx).value = f"=SUM({col_letter}2:{col_letter}{sum_row-1})"
        ws.cell(row=sum_row, column=col_idx).number_format = '$#,##0.00'
        ws.cell(row=sum_row, column=col_idx).font = sum_font

    ws.cell(row=sum_row, column=7, value=f"=SUM(G2:G{sum_row-1})").font = sum_font
    ws.cell(row=sum_row, column=7).number_format = '#,##0'

    auto_fit_columns(ws)
    ws.freeze_panes = "A2"

    # --- Dashboard sheet ---
    ws2 = wb.create_sheet("Dashboard")
    ws2.append(["Invoice Processing Dashboard", "", ""])
    ws2.merge_cells("A1:C1")

    totals = [r[10] for r in rows_data]
    status_counts = {}
    vendor_totals = {}
    for r in rows_data:
        status_counts[r[4]] = status_counts.get(r[4], 0) + 1
        vendor_totals[r[1]] = vendor_totals.get(r[1], 0) + r[10]

    status_amounts = {}
    for r in rows_data:
        status_amounts[r[4]] = status_amounts.get(r[4], 0) + r[10]

    ws2.append(["Total Invoices Processed", len(rows_data), ""])
    ws2.append(["Total Amount (Grand Total)", f"${sum(totals):,.2f}", ""])
    ws2.append(["Average Invoice Amount", f"${sum(totals)/len(totals):,.2f}", ""])
    ws2.append(["Date Range", f"{rows_data[0][2]} to {rows_data[-1][2]}", ""])
    ws2.append(["", "", ""])
    ws2.append(["Status", "Count", "Total Amount"])
    for st in ["Paid", "Pending", "Overdue"]:
        c = status_counts.get(st, 0)
        a = status_amounts.get(st, 0)
        ws2.append([st, c, f"${a:,.2f}"])
    ws2.append(["", "", ""])
    ws2.append(["Payment Method", "Count", ""])
    pm_counts = {}
    for r in rows_data:
        pm_counts[r[11]] = pm_counts.get(r[11], 0) + 1
    for pm, c in sorted(pm_counts.items(), key=lambda x: -x[1]):
        ws2.append([pm, c, f"{c/len(rows_data)*100:.1f}%"])
    ws2.append(["", "", ""])
    ws2.append(["Top 10 Vendors by Total Amount", "Invoices", "Total Amount"])
    sorted_vendors = sorted(vendor_totals.items(), key=lambda x: -x[1])[:10]
    vendor_inv_count = {}
    for r in rows_data:
        vendor_inv_count[r[1]] = vendor_inv_count.get(r[1], 0) + 1
    for v, amt in sorted_vendors:
        ws2.append([v, vendor_inv_count[v], f"${amt:,.2f}"])

    style_summary_sheet(ws2, "C44900")

    return wb


# ===========================================================================
# MAIN
# ===========================================================================

if __name__ == "__main__":
    import os

    output_dir = r"c:\Users\rafca\OneDrive\Desktop\Toxic or Nah\ryan_cole\portfolio"
    os.makedirs(output_dir, exist_ok=True)

    random.seed(42)  # Reproducible

    print("Generating Sample 1: E-commerce Product Scraping...")
    wb1 = generate_ecommerce()
    path1 = os.path.join(output_dir, "portfolio_ecommerce_scraping.xlsx")
    wb1.save(path1)
    print(f"  Saved: {path1}")

    print("Generating Sample 2: B2B Dental Leads - Texas...")
    wb2 = generate_dental_leads()
    path2 = os.path.join(output_dir, "portfolio_dental_leads_texas.xlsx")
    wb2.save(path2)
    print(f"  Saved: {path2}")

    print("Generating Sample 3: Invoice Processing...")
    wb3 = generate_invoice_processing()
    path3 = os.path.join(output_dir, "portfolio_invoice_processing.xlsx")
    wb3.save(path3)
    print(f"  Saved: {path3}")

    print("\nAll 3 portfolio samples generated successfully!")
