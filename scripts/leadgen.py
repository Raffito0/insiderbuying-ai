#!/usr/bin/env python3
"""Lead Generation Script - uses SerpApi Google Maps"""
import os, sys

os.system("pip install --break-system-packages google-search-results openpyxl -q 2>/dev/null")

from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

query = sys.argv[1] if len(sys.argv) > 1 else "Italian restaurants in Manhattan New York"
count = int(sys.argv[2]) if len(sys.argv) > 2 else 20
output = sys.argv[3] if len(sys.argv) > 3 else "/root/.openclaw/workspace/leadgen_output.xlsx"

params = {
    "engine": "google_maps",
    "q": query,
    "api_key": os.environ.get("SERPAPI_KEY"),
    "num": count
}

search = GoogleSearch(params)
results = search.get_dict()
places = results.get("local_results", [])[:count]

wb = Workbook()
ws = wb.active
ws.title = "Lead Results"

headers = ["Name", "Address", "Phone", "Website", "Rating"]
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for i, place in enumerate(places, 2):
    ws.cell(row=i, column=1, value=place.get("title", ""))
    ws.cell(row=i, column=2, value=place.get("address", ""))
    ws.cell(row=i, column=3, value=place.get("phone", ""))
    ws.cell(row=i, column=4, value=place.get("website", ""))
    ws.cell(row=i, column=5, value=place.get("rating", ""))

for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

wb.save(output)
print(f"Saved {len(places)} results to {output}")
